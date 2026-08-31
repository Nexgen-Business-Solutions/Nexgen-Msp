import datetime
import re

from nexgen_msp.utils.errors import ValidationError

EXCLUDED_COLUMNS = ("user", "password")

COLUMN_INDEX = {
    "full_name": 0,
    "username": 1,
    "password": 2,
    "email": 3,
    "company": 4,
    "department": 5,
    "ad_created": 6,
    "ad_disabled": 7,
    "parallels": 8,
    "nextcloud": 9,
    "nextcloud_activated": 10,
    "nextcloud_disabled": 11,
    "sophos": 12,
    "hostname": 13,
    "device_created": 14,
    "device_disabled": 15,
    "mac_wifi": 16,
    "mac_lan": 17,
    "mac_extra": 18,
    "device_type": 19,
    "remarks": 20,
}


EXPECTED_HEADERS = ("full name", "user", "password", "email")

# Columns the sheet may or may not carry. Each is found by its header wherever it sits,
# so a file written to the older layout imports exactly as before. Add or reword an entry
# here if a workbook names a column differently — anything not matched is simply left empty.
OPTIONAL_COLUMNS = {
    "last_billed_on": (
        "dernier date de facturation",
        "derniere date de facturation",
        "dernière date de facturation",
        "derniere facturation",
        "dernière facturation",
        "last billing",
        "last billed",
        "last billing date",
    ),
    "covered_until": (
        "periode facturation",
        "période facturation",
        "billing period",
        "covered until",
    ),
    "has_export_link": ("has export link",),
    "export_link": ("link",),
}

# "jusqu'a 31-07-2026" says the same thing as a bare date; "N/A" and "never invoiced" say
# there is nothing to record.
COVERED_PREFIXES = ("jusqu'a", "jusqu'à", "jusquà", "until", "up to")

NOT_COVERED_TOKENS = {"n/a", "na", "never invoiced", "never", "-", ""}

# the optional columns are appended after the fixed block by load_rows, in this order
OPTIONAL_INDEX = {
    key: len(COLUMN_INDEX) + offset for offset, key in enumerate(OPTIONAL_COLUMNS)
}

BLANK_TOKENS = {"", "n/a", "na", "none", "-", "null"}

DEVICE_TYPE_MAP = {
    "pc": "PC",
    "laptop": "Laptop",
    "mini pc": "Mini PC",
    "minipc": "Mini PC",
    "mac": "Mac",
    "tablette": "Tablet",
    "tablet": "Tablet",
    "server": "Server",
    "serveur": "Server",
    "firewall": "Firewall",
    "vm": "VM",
}

MAC_PATTERN = re.compile(r"^[0-9A-F]{2}([:-])(?:[0-9A-F]{2}\1){4}[0-9A-F]{2}$")

BIDI_PATTERN = re.compile(r"[‎‏‪-‮⁦-⁩]")

TEXT_DATE_FORMATS = ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d")

ACTIVE_TOKENS = {"active", "actif"}


def is_blank(value):
    if value is None:
        return True
    if isinstance(value, datetime.datetime):
        return False
    return str(value).strip().lower() in BLANK_TOKENS


def as_text(value):
    if is_blank(value):
        return None
    return " ".join(BIDI_PATTERN.sub("", str(value)).split())


def as_date(value):
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value

    text = as_text(value)
    if not text:
        return None

    for fmt in TEXT_DATE_FORMATS:
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def is_marked_active(value):
    text = as_text(value)
    return bool(text) and text.strip().lower() in ACTIVE_TOKENS


def resolve_service(assigned, lifecycle_value, start_date, own_lifecycle=False):
    """Turn an Excel service flag and its lifecycle column into an assignment state."""
    end_date = as_date(lifecycle_value)
    marked_active = is_marked_active(lifecycle_value)

    if own_lifecycle and not assigned:
        assigned = bool(end_date) or marked_active

    if not assigned:
        return {"assigned": False, "status": None, "start": None, "end": None, "inconsistent_start": False}

    inconsistent = bool(start_date and end_date and end_date < start_date)

    return {
        "assigned": True,
        "status": "Ended" if end_date else "Active",
        "start": None if inconsistent else start_date,
        "end": end_date,
        "inconsistent_start": inconsistent,
    }


def is_yes(value):
    return not is_blank(value) and str(value).strip().lower() in ("yes", "oui", "y", "1", "true")


def normalize_mac(value):
    if is_blank(value):
        return None
    candidate = str(value).strip().upper().replace(".", ":")
    return candidate if MAC_PATTERN.match(candidate) else False


def normalize_device_type(value):
    text = as_text(value)
    if not text:
        return None
    return DEVICE_TYPE_MAP.get(text.strip().lower())


def load_rows(file_path):
    try:
        import openpyxl
    except ImportError:
        raise ValidationError("openpyxl is not installed on the server.", "INTERNAL_ERROR", 500)

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        raise ValidationError(f"Cannot read the workbook: {e}", "INVALID_FILE")

    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise ValidationError("The workbook is empty.", "INVALID_FILE")

    header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]

    for expected in EXPECTED_HEADERS:
        if expected not in header:
            raise ValidationError(
                f"Column '{expected}' is missing from the workbook.", "MISSING_COLUMN"
            )

    # a header may be worded loosely, so match on how it starts as well as in full
    def locate(names):
        for name in names:
            if name in header:
                return header.index(name)

        for index, cell in enumerate(header):
            if any(cell.startswith(name) for name in names if cell):
                return index

        return None

    found = {key: locate(names) for key, names in OPTIONAL_COLUMNS.items()}

    # The rest of the sheet is read by position, so the optional columns are lifted out of
    # the row and appended at the end — they can then sit anywhere without shifting the
    # columns that follow them.
    lifted = sorted((index for index in found.values() if index is not None), reverse=True)
    keep = len(COLUMN_INDEX)
    trimmed = []

    for row in rows[1:]:
        if not any(not is_blank(cell) for cell in row):
            continue

        cells = list(row)
        extras = {key: (cells[index] if index is not None and index < len(cells) else None)
                  for key, index in found.items()}

        for index in lifted:
            if index < len(cells):
                cells.pop(index)

        block = cells[:keep]
        block.extend([None] * (keep - len(block)))
        block.extend(extras[key] for key in OPTIONAL_COLUMNS)
        trimmed.append(tuple(block))

    return trimmed


def read(row, key):
    index = OPTIONAL_INDEX[key] if key in OPTIONAL_INDEX else COLUMN_INDEX[key]
    return row[index] if index < len(row) else None


def as_covered_until(value):
    """A coverage cell: a date, a "jusqu'a <date>" phrase, or a way of saying "none"."""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return as_date(value)

    text = (as_text(value) or "").strip()

    if text.lower() in NOT_COVERED_TOKENS:
        return None

    lowered = text.lower()

    for prefix in COVERED_PREFIXES:
        if lowered.startswith(prefix):
            text = text[len(prefix) :].strip()
            break

    return as_date(text)


def parse_row(row, row_number):
    record = {
        "row_number": row_number,
        "full_name": as_text(read(row, "full_name")),
        # the account name their licences are issued against — not a credential, unlike
        # the password column beside it, which is never read
        "username": as_text(read(row, "username")),
        "email": as_text(read(row, "email")),
        "company": as_text(read(row, "company")),
        "department": as_text(read(row, "department")),
        "ad_created": as_date(read(row, "ad_created")),
        "ad_disabled": as_date(read(row, "ad_disabled")),
        "ad_marked_active": is_marked_active(read(row, "ad_disabled")),
        "hostname": as_text(read(row, "hostname")),
        "device_type": normalize_device_type(read(row, "device_type")),
        "device_type_raw": as_text(read(row, "device_type")),
        "device_created": as_date(read(row, "device_created")),
        "device_disabled": as_date(read(row, "device_disabled")),
        "remarks": as_text(read(row, "remarks")),
        # absent from older sheets, so these stay empty rather than being invented
        "last_billed_on": as_date(read(row, "last_billed_on")),
        "covered_until": as_covered_until(read(row, "covered_until")),
        "has_export_link": is_yes(read(row, "has_export_link")),
        "export_link": as_text(read(row, "export_link")),
        "services": {
            "parallels": resolve_service(
                is_yes(read(row, "parallels")),
                read(row, "ad_disabled"),
                as_date(read(row, "ad_created")),
            ),
            "nextcloud": resolve_service(
                is_yes(read(row, "nextcloud")),
                read(row, "nextcloud_disabled"),
                as_date(read(row, "nextcloud_activated")),
                own_lifecycle=True,
            ),
            "sophos": resolve_service(
                not is_blank(read(row, "sophos")),
                read(row, "device_disabled"),
                as_date(read(row, "device_created")),
            ),
        },
        "macs": [],
        "invalid_macs": [],
    }

    if record["hostname"]:
        record["hostname"] = record["hostname"].upper()

    for key, interface_type in (
        ("mac_wifi", "Wi-Fi"),
        ("mac_lan", "LAN"),
        ("mac_extra", "Extra"),
    ):
        raw = read(row, key)
        if is_blank(raw):
            continue
        mac = normalize_mac(raw)
        if mac is False:
            record["invalid_macs"].append(str(raw).strip())
        else:
            record["macs"].append({"interface_type": interface_type, "mac_address": mac})

    return record
