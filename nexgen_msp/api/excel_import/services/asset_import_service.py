"""The second sheet: a hostname, and the two facts we still lack about it.

The hostname is the key. It says which machine carries the serial number, and — through
whoever holds that machine — which person the username belongs to.

Column names are matched on what they contain rather than on an exact spelling, because
the file comes from elsewhere and its headers are worded loosely.
"""

import frappe

from nexgen_msp.utils.errors import ValidationError

# what a header must contain for us to recognise it; the first match wins
COLUMNS = {
    "hostname": ("hostname", "host name", "host", "machine", "poste", "pc", "device"),
    "serial_number": ("serial", "serie", "série", "sn", "s/n", "numero de serie"),
    "username": ("username", "user name", "login", "utilisateur", "compte", "account"),
}


class AssetImportService:
    @staticmethod
    def import_assets(file_url=None, dry_run=1, fill_blanks_only=1):
        """Read the sheet, apply what it says, and account for every line."""
        from nexgen_msp.api.excel_import.services.excel_import_service import ExcelImportService

        dry_run = frappe.utils.cint(dry_run)
        fill_blanks_only = frappe.utils.cint(fill_blanks_only)

        if not frappe.has_permission("MSP Managed Device", "write"):
            raise ValidationError(
                "You are not allowed to import devices.", "PERMISSION_DENIED", 403
            )

        rows = AssetImportService._read(ExcelImportService._resolve_path(file_url))

        report = {
            "dry_run": dry_run,
            "rows_read": len(rows),
            "updated": {"serial_numbers": 0, "usernames": 0},
            "skipped": {
                "no_hostname": 0,
                "unknown_hostname": 0,
                "ambiguous_hostname": 0,
                "nothing_to_write": 0,
                "already_filled": 0,
                "no_holder": 0,
                "serial_taken": 0,
            },
            "exceptions": [],
        }

        try:
            for row in rows:
                AssetImportService._apply(row, report, fill_blanks_only)

            if dry_run:
                frappe.db.rollback()
            else:
                frappe.db.commit()
        except Exception:
            frappe.db.rollback()
            raise

        return report

    @staticmethod
    def _apply(row, report, fill_blanks_only):
        hostname = (row.get("hostname") or "").strip().upper()
        serial = (row.get("serial_number") or "").strip()
        username = (row.get("username") or "").strip()
        number = row["row_number"]

        if not hostname:
            report["skipped"]["no_hostname"] += 1
            return

        if not serial and not username:
            report["skipped"]["nothing_to_write"] += 1
            return

        devices = frappe.get_all(
            "MSP Managed Device",
            filters={"hostname": hostname},
            fields=["name", "customer", "serial_number", "assigned_client_user"],
        )

        if not devices:
            report["skipped"]["unknown_hostname"] += 1
            report["exceptions"].append(
                {"row": number, "reason": f"No device here is called {hostname}."}
            )
            return

        if len(devices) > 1:
            # the same name can exist at two customers, so the sheet cannot say which one
            report["skipped"]["ambiguous_hostname"] += 1
            report["exceptions"].append(
                {
                    "row": number,
                    "reason": f"{hostname} exists at {len(devices)} customers "
                    f"({', '.join(d.customer for d in devices)}) — nothing was written.",
                }
            )
            return

        device = devices[0]

        if serial:
            AssetImportService._write_serial(device, serial, number, report, fill_blanks_only)

        if username:
            AssetImportService._write_username(device, username, number, report, fill_blanks_only)

    @staticmethod
    def _write_serial(device, serial, number, report, fill_blanks_only):
        if device.serial_number and fill_blanks_only:
            if device.serial_number.strip() != serial:
                report["skipped"]["already_filled"] += 1
                report["exceptions"].append(
                    {
                        "row": number,
                        "reason": f"{device.name} already carries serial "
                        f"{device.serial_number} — the sheet says {serial}, kept as it is.",
                    }
                )
            return

        twin = frappe.db.get_value(
            "MSP Managed Device",
            {"serial_number": serial, "name": ("!=", device.name)},
            ["name", "hostname"],
            as_dict=True,
        )

        if twin:
            report["skipped"]["serial_taken"] += 1
            report["exceptions"].append(
                {
                    "row": number,
                    "reason": f"Serial {serial} is already on {twin.hostname} ({twin.name}).",
                }
            )
            return

        frappe.db.set_value("MSP Managed Device", device.name, "serial_number", serial)
        report["updated"]["serial_numbers"] += 1

    @staticmethod
    def _write_username(device, username, number, report, fill_blanks_only):
        """The username belongs to the person holding the machine, not to the machine."""
        holder = device.assigned_client_user

        if not holder:
            report["skipped"]["no_holder"] += 1
            report["exceptions"].append(
                {
                    "row": number,
                    "reason": f"{device.hostname} is held by nobody, so there is no one to "
                    f"give the username {username} to.",
                }
            )
            return

        held = (frappe.db.get_value("MSP Client User", holder, "username") or "").strip()

        if held and fill_blanks_only:
            if held != username:
                report["skipped"]["already_filled"] += 1
                report["exceptions"].append(
                    {
                        "row": number,
                        "reason": f"{holder} already has username {held} — the sheet says "
                        f"{username}, kept as it is.",
                    }
                )
            return

        frappe.db.set_value("MSP Client User", holder, "username", username)
        report["updated"]["usernames"] += 1

    @staticmethod
    def _read(path):
        try:
            import openpyxl
        except ImportError:
            raise ValidationError("openpyxl is not installed on the server.", "INTERNAL_ERROR", 500)

        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            raise ValidationError(f"Cannot read the workbook: {e}", "INVALID_FILE")

        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            raise ValidationError("The workbook is empty.", "INVALID_FILE")

        header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        found = AssetImportService._locate(header)

        if "hostname" not in found:
            raise ValidationError(
                "No hostname column found. The sheet needs one column naming the machine, "
                "and at least one of serial number or username.",
                "MISSING_COLUMN",
            )

        if "serial_number" not in found and "username" not in found:
            raise ValidationError(
                "Found the hostname column but neither a serial number nor a username "
                "column — there would be nothing to write.",
                "MISSING_COLUMN",
            )

        out = []

        for number, row in enumerate(rows[1:], start=2):
            if not any(cell not in (None, "") for cell in row):
                continue

            record = {"row_number": number}

            for key, index in found.items():
                value = row[index] if index < len(row) else None
                record[key] = "" if value is None else str(value).strip()

            out.append(record)

        return out

    @staticmethod
    def _locate(header):
        """Match each column on what its header contains, not on an exact spelling."""
        found = {}
        used = set()

        for key, needles in COLUMNS.items():
            for index, cell in enumerate(header):
                if index in used or not cell:
                    continue

                if any(needle in cell for needle in needles):
                    found[key] = index
                    used.add(index)
                    break

        return found

    @staticmethod
    def describe(file_url=None):
        """What we recognised in the sheet, so the operator can check before importing."""
        from nexgen_msp.api.excel_import.services.excel_import_service import ExcelImportService

        try:
            import openpyxl
        except ImportError:
            raise ValidationError("openpyxl is not installed on the server.", "INTERNAL_ERROR", 500)

        workbook = openpyxl.load_workbook(
            ExcelImportService._resolve_path(file_url), read_only=True, data_only=True
        )
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True, max_row=1))
        header = [str(cell).strip() if cell is not None else "" for cell in (rows[0] if rows else [])]
        found = AssetImportService._locate([cell.lower() for cell in header])

        return {
            "headers": header,
            "recognised": {key: header[index] for key, index in found.items()},
            "missing": [key for key in COLUMNS if key not in found],
        }
