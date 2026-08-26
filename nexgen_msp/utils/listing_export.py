import io

import frappe
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="2F82F8")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BODY = Font(size=10)


def workbook(title, columns, rows):
	"""One sheet, one row per record, the columns the screen is showing.

	The screen decides what matters; this only writes it down. Anything the caller did not
	ask for stays out, so the file reads like the page rather than like the table.
	"""
	book = Workbook()
	sheet = book.active
	sheet.title = title[:31]

	for index, (_key, label) in enumerate(columns, start=1):
		cell = sheet.cell(row=1, column=index, value=label)
		cell.fill = HEADER_FILL
		cell.font = HEADER_FONT
		cell.alignment = Alignment(horizontal="center")
		sheet.column_dimensions[get_column_letter(index)].width = max(len(label) + 4, 14)

	for line, row in enumerate(rows, start=2):
		for index, (key, _label) in enumerate(columns, start=1):
			cell = sheet.cell(row=line, column=index, value=_as_text(row.get(key)))
			cell.font = BODY

	sheet.freeze_panes = "A2"

	stream = io.BytesIO()
	book.save(stream)
	return stream.getvalue()


def _as_text(value):
	if value is None:
		return ""

	if isinstance(value, (int, float)):
		return value

	return str(value)


def respond(filename, title, columns, rows):
	"""Hand the sheet to the browser as a download."""
	frappe.local.response.filename = filename
	frappe.local.response.filecontent = workbook(title, columns, rows)
	frappe.local.response.type = "download"
