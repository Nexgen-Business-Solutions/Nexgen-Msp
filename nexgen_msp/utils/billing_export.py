import io

import frappe
from frappe.utils import flt, getdate
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13)
BOLD = Font(bold=True, size=10)
BODY = Font(size=10)
THIN = Side(style="thin", color="D0D7DE")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DETAIL_COLUMNS = (
    ("service_name", "Service", 26),
    ("count", "#", 5),
    ("employee_name", "Employee Name", 26),
    ("email", "Email", 30),
    ("hostname", "Host name", 20),
    ("company", "Company", 16),
    ("department", "Department", 20),
    ("creation_date", "Creation Date", 14),
    ("reference", "Reference", 22),
    ("status", "Status", 14),
    ("monthly", "Monthly", 12),
    ("months", "Months", 9),
    ("total", "Total", 14),
    ("comments", "Comments", 34),
)


def money_format(currency):
    """Excel wants the symbol in the cell format, so a total still reads as money."""
    symbol = frappe.db.get_value("Currency", currency, "symbol") if currency else None
    suffix = f'" {symbol or currency}"' if (symbol or currency) else ""

    return f"#,##0.00{suffix}"


MONTHS = "0.0#"


def _write_header(sheet, row, labels):
    for index, label in enumerate(labels, start=1):
        cell = sheet.cell(row=row, column=index, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[row].height = 26
    return row + 1


def _as_text(value):
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return getdate(value).strftime("%Y-%m-%d")
    return value


def breakdown_workbook(data):
    """The supporting detail file: a summary block, then one table per service."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (data["period_label"] or "Breakdown")[:31]

    currency = data.get("currency") or ""
    money = money_format(currency)

    sheet.cell(row=1, column=1, value=f"{data['customer']} — {data['period_label']}").font = (
        TITLE_FONT
    )
    sheet.cell(
        row=2,
        column=1,
        value=f"{data['period_start']} to {data['period_end']}"
        + (f" · invoice {data['invoice']}" if data.get("invoice") else ""),
    ).font = BODY

    row = 4
    row = _write_header(sheet, row, ["Company Name", "Service", "Months", data["period_label"]])

    for entry in data["summary"]:
        sheet.cell(row=row, column=1, value=entry["company"]).font = BODY
        sheet.cell(row=row, column=2, value=entry["service"]).font = BODY
        months = sheet.cell(row=row, column=3, value=entry["months"])
        months.font = BODY
        months.number_format = MONTHS
        months.alignment = Alignment(horizontal="center")
        amount = sheet.cell(row=row, column=4, value=entry["amount"])
        amount.font = BODY
        amount.number_format = money
        for column in range(1, 5):
            sheet.cell(row=row, column=column).border = BOX
        row += 1

    sheet.cell(row=row, column=2, value=f"Total invoice {data['period_label']}").font = BOLD
    total = sheet.cell(row=row, column=4, value=data["total_amount"])
    total.font = BOLD
    total.number_format = money
    for column in range(1, 5):
        sheet.cell(row=row, column=column).border = BOX

    row += 3

    for block in data["blocks"]:
        sheet.cell(row=row, column=1, value=block["service_name"]).font = TITLE_FONT
        row += 1
        row = _write_header(sheet, row, [label for _key, label, _width in DETAIL_COLUMNS])

        for entry in block["rows"]:
            values = dict(entry)
            values["service_name"] = block["service_name"]

            for index, (key, _label, _width) in enumerate(DETAIL_COLUMNS, start=1):
                cell = sheet.cell(row=row, column=index, value=_as_text(values.get(key)))
                cell.font = BODY
                cell.border = BOX
                if key in ("monthly", "total"):
                    cell.number_format = money
                if key == "months":
                    cell.number_format = MONTHS
                if key in ("count", "months"):
                    cell.alignment = Alignment(horizontal="center")
            row += 1

        sheet.cell(row=row, column=10, value="Subtotal").font = BOLD
        months = sheet.cell(row=row, column=12, value=flt(block["months"], 2))
        months.font = BOLD
        months.number_format = MONTHS
        months.alignment = Alignment(horizontal="center")
        subtotal = sheet.cell(row=row, column=13, value=flt(block["total"], 2))
        subtotal.font = BOLD
        subtotal.number_format = money
        for column in range(1, len(DETAIL_COLUMNS) + 1):
            sheet.cell(row=row, column=column).border = BOX

        row += 3

    for index, (_key, _label, width) in enumerate(DETAIL_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A5"

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def attach_breakdown(run_name):
    """Build the breakdown and keep it on the run, so the same file is what everyone sees."""
    from nexgen_msp.api.internal.services.billing_service import BillingService

    data = BillingService.breakdown(run_name)
    content = breakdown_workbook(data)
    filename = f"Breakdown-{data['customer']}-{data['period_label']}.xlsx".replace(" ", "-")

    existing = frappe.db.get_value(
        "File",
        {"attached_to_doctype": "Billing Run", "attached_to_name": run_name, "file_name": filename},
        "name",
    )

    if existing:
        frappe.delete_doc("File", existing, ignore_permissions=True, force=True)

    file = frappe.get_doc(
        {
            "doctype": "File",
            "file_name": filename,
            "attached_to_doctype": "Billing Run",
            "attached_to_name": run_name,
            "is_private": 1,
            "content": content,
        }
    ).insert(ignore_permissions=True)

    return {"file_url": file.file_url, "file_name": file.file_name}
